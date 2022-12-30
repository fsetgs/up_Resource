from celery import shared_task
from checkControl.models import *
from django.db import close_old_connections
from Resource.settings import *
from openpyxl import Workbook
import requests,json,time
from django.shortcuts import render,redirect,HttpResponse
from django.http import JsonResponse
@shared_task
def addcheck(checkCompanyName,id):
    close_old_connections()
    company_obj = company.objects.get(name=checkCompanyName)
    res_obj = resource.objects.filter(company=company_obj)
    for res in res_obj:
        check_resource.objects.create(t_check_id=id,t_resource=res,is_checked="盘点中")

@shared_task
def stopcheck(id):
    close_old_connections()
    relate_obj = check_resource.objects.filter(t_check_id=id,is_checked="盘点中")
    for obj in relate_obj:
        obj.is_checked = "未盘点"
        obj.save()

@shared_task
def deletecheck(id):
    close_old_connections()
    relate_obj = check_resource.objects.filter(t_check_id=id)
    relate_obj.delete()

# @shared_task
# def check_out_file(id,status):
#     close_old_connections()
#     if status == "all":
#         relate_obj = check_resource.objects.filter(t_check_id=id)
#     else:
#         relate_obj = check_resource.objects.filter(t_check_id=id,is_checked=status)
#     upfile_url = MEDIA_ROOT + '/tempUpfile/' + relate_obj[0].t_check.name + ".xlsx"
#     wb = Workbook()
#     wb.create_sheet(index=0,title="Sheet1")
#     ws = wb.active
#     ws['A1'] = '所属盘点'
#     ws['B1'] = '资产编码'
#     ws['C1'] = '资产名称'
#     ws['D1'] = '盘点状态'
#     wb.save(upfile_url)
    
    
