from celery import shared_task
from .models import *
import openpyxl
from Resource.settings import *
from dateutil import rrule
import datetime
from openpyxl import Workbook
from django.db import close_old_connections
@shared_task
def updata(files):
    close_old_connections()
    # for file in files:
    #     up_url = MEDIA_ROOT + '/tempUpfile/' + file.name
    #     print(up_url)
    #     with open(up_url,'wb+') as f:
    #         for chunk in file.chunks():
    #             f.write(chunk)
    #     # 开始读取xlsx文件
    #     workbook = openpyxl.load_workbook(up_url,data_only=True)
    #     sheet = workbook['Sheet1']
    #     rows = sheet.max_row #获取最大行
    #     # 按行获取值
    #     for row_obj in sheet.iter_rows(min_row=2,max_row=rows,min_col=3,max_col=30):
    #         if row_obj[0] is None:
    #             continue
    #         company_name = row_obj[4].value
    #         # 判断公司是否存在
    #         try:
    #             company_obj = company.objects.get(name=company_name)
    #             company_id = company_obj.id
    #         except:
    #             print("公司跳出",row_obj[4].value)
    #             continue
    #         # 判断部门是否存在
    #         try:
    #             department_name = row_obj[5].value
    #             department_obj = department.objects.get(name=department_name,company_id=company_id)
    #             print(department_obj,"1111111111")
    #         except Exception as e:
    #             print("部门跳出",company_id,row_obj[5].value,e)
    #             continue
    #         try:
    #             type_obj = resourceType.objects.get(code=row_obj[1].value)
    #         except:
    #             print("类型跳出",row_obj[1].value)
    #             continue
            
    #         name = row_obj[0].value
    #         resource_from = row_obj[2].value
    #         resource_status = row_obj[3].value
    #         localtion_area = row_obj[6].value
    #         location = row_obj[7].value
    #         duty = row_obj[8].value
    #         borrow_department = row_obj[9].value
    #         borrow_user = row_obj[10].value
    #         borrow_time = datetime.datetime.strftime(row_obj[11].value,"%Y-%m-%d")
    #         return_time = datetime.datetime.strftime(row_obj[12].value,"%Y-%m-%d")
    #         storage_time = datetime.datetime.strftime(row_obj[13].value,"%Y-%m-%d")
    #         buy_time = datetime.datetime.strftime(row_obj[14].value,"%Y-%m-%d")
    #         price = row_obj[15].value
    #         depreciation_period = row_obj[16].value
    #         residuals_rate = row_obj[17].value
    #         resource_residuals = row_obj[18].value #资产残值
    #         month_depreciation = row_obj[19].value #月折旧额
    #         total_depreciation = row_obj[20].value #累计折旧
    #         net_value = row_obj[21].value #资产净值
    #         specifications = row_obj[22].value
    #         units = row_obj[23].value
    #         provider = row_obj[24].value
    #         sn = row_obj[25].value
    #         mac = row_obj[26].value
    #         comment = row_obj[27].value
    #         if resource_residuals == "" or resource_residuals is None:
    #             resource_residuals = str(int(int(price)*int(residuals_rate)/100))
    #         if month_depreciation == "" or month_depreciation is None:
    #             month_depreciation = round(float(int(price)*(1-int(residuals_rate)/100)/int(depreciation_period)),2)
    #         if total_depreciation == "" or total_depreciation is None:
    #             storage_date = datetime.datetime.strptime(storage_time,"%Y-%m-%d")
    #             months = rrule.rrule(rrule.MONTHLY,dtstart=storage_date,until=datetime.datetime.now()).count()
    #             if months <= int(depreciation_period):
    #                 total_depreciation = str(months*int(month_depreciation))
    #         if net_value == "" or net_value is None:
    #             net_value = str(int(price)-int(total_depreciation))

    #         resource_info_obj = resourceInfo.objects.create(storage_time=storage_time,buy_time=buy_time,resource_price=price,depreciation_period=depreciation_period,
    #         residuals_rate=residuals_rate,resource_residuals=resource_residuals,units=units,provider=provider,mac=mac,sn=sn,comment=comment,
    #         specifications=specifications)

    #         res_obj = resource.objects.create(name=name,location=location,location_area=localtion_area,duty=duty,user=borrow_user,borrow_department=borrow_department,
    #         company=company_obj,resource_type=type_obj,resource_from=resource_from,resource_status=resource_status,borrow_time=borrow_time,
    #         return_time=return_time,detail_info=resource_info_obj)

    #         res_num = res_obj.id
    #         if len(str(res_num)) < 2:
    #             num = "00" + str(res_num)
    #         elif 1 < len(str(res_num)) < 3:
    #             num = "0" + str(res_num)
    #         elif 2 < len(str(res_num)):
    #             num = str(res_num)
    #         # 创建编号 公司编码+资产类别编码+部门编码+采购年月(1612表16年12月)+顺序号
    #         resource_code = str(company_obj.code) + str(row_obj[1].value) + str(department_obj.code) + str(buy_time[2:4])+ str(buy_time[5:7]) + num
    #         res_obj.code = resource_code
    #         res_obj.save()
    #     os.remove(up_url)
    row_num = 2
    error_line =  2
    row_list = []
    code_list = []
    for file in files:
        up_url = MEDIA_ROOT + '/tempUpfile/' + file.name
        print(up_url)
        with open(up_url,'wb+') as f:
            for chunk in file.chunks():
                f.write(chunk)
        # 开始读取xlsx文件
        workbook = openpyxl.load_workbook(up_url,data_only=True)
        sheet = workbook['Sheet1']
        rows = sheet.max_row #获取最大行


        # 将录入失败的数据综合出来
        upfile_url = MEDIA_ROOT + '/samedata/' + file.name
        wb = Workbook()
        wb.create_sheet(index=0,title="Sheet1")
        ws = wb.active
        ws['A1'] = '未录入的资产编码'
        ws['B1'] = '问题数据所在行'
        ws['C1'] = '问题可能原因'

        # 按行获取值
        for row_obj in sheet.iter_rows(min_row=2,max_row=rows,min_col=2,max_col=30):
            try:
                code = row_obj[0].value
                company_name = row_obj[5].value
                company_obj = company.objects.get(name=company_name)
                company_id = company_obj.id
                department_name = row_obj[6].value
                department_obj = department.objects.get(name=department_name,company_id=company_id)
                type_obj = resourceType.objects.get(code=code[2:4])
                name = row_obj[1].value
                resource_from = row_obj[3].value
                resource_status = row_obj[4].value
                localtion_area = row_obj[7].value
                location = row_obj[8].value
                duty = row_obj[9].value
                borrow_department = row_obj[10].value
                borrow_user = row_obj[11].value
                borrow_time = row_obj[12].value
                return_time = row_obj[13].value
                storage_time = row_obj[14].value
                buy_time = row_obj[15].value
                price = row_obj[16].value
                depreciation_period = row_obj[17].value
                residuals_rate = row_obj[18].value
                resource_residuals = row_obj[19].value #资产残值
                month_depreciation = row_obj[20].value #月折旧额
                total_depreciation = row_obj[21].value #累计折旧
                net_value = row_obj[22].value #资产净值
                specifications = row_obj[23].value
                units = row_obj[24].value
                provider = row_obj[25].value
                sn = row_obj[26].value
                mac = row_obj[27].value
                comment = row_obj[28].value
                res_obj = resource.objects.filter(code=code)
                if len(res_obj) == 0:
                    resource_info_obj = resourceInfo.objects.create(storage_time=storage_time,buy_time=buy_time,resource_price=price,depreciation_period=depreciation_period,
                    residuals_rate=residuals_rate,resource_residuals=resource_residuals,units=units,provider=provider,mac=mac,sn=sn,comment=comment,
                    specifications=specifications)

                    resource.objects.create(code=code,name=name,location=location,location_area=localtion_area,duty=duty,user=borrow_user,borrow_department=borrow_department,
                    company=company_obj,resource_type=type_obj,resource_from=resource_from,resource_status=resource_status,borrow_time=borrow_time,
                    return_time=return_time,detail_info=resource_info_obj,department=department_obj)
                    row_num += 1
                else:
                    # 综合问题数据
                    ws[f'A{error_line}'] = code
                    ws[f'B{error_line}'] = row_num
                    ws[f'C{error_line}'] = "1.录入时文件中有两个相同的编码 2.该条数据已存在后台中"
                    error_line += 1

                    code_list.append(code)
                    row_list.append(row_num)
                    row_num += 1
            except Exception as e:
                code = row_obj[0].value
                
                # 综合问题数据
                ws[f'A{error_line}'] = code
                ws[f'B{error_line}'] = row_num
                ws[f'C{error_line}'] = "1.所属公司与后台录入不匹配/后台未录入 2.所属部门未在后台所属公司中录入 3.资产编码3-4位大于19 4.此条数据格式有问题"
                error_line += 1

                # code_list.append(code)
                row_list.append(row_num)
                print("未录入行：",row_num,e)
                row_num += 1
                continue
        print("总条数",row_num-2)
        print("资产编码冲突的",code_list)
        print("未录入的",row_list)

        if error_line > 2:
            wb.save(upfile_url)
        os.remove(up_url)
